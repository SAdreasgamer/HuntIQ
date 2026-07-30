"""
HuntIQ — Excel Report Generator Service.

Generates professional, multi-worksheet Excel workbooks (.xlsx) using openpyxl:
- Executive Summary KPI Dashboard
- Top AI Matched Jobs (>= 80% match score) with conditional formatting
- Full Discovered Jobs Inventory with auto-filter headers
- Application Kanban Tracker & Recruiter Directory
- Tech Stack Skill Demand Frequency Breakdown
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.analytics.engine import AnalyticsEngineService
from app.core.logging import get_logger
from app.repositories.application import ApplicationRepository
from app.repositories.job import JobRepository
from app.repositories.user import UserRepository

logger = get_logger(__name__)

# Styling Constants
COLOR_HEADER_BG = "1F4E79"       # Deep Navy
COLOR_HEADER_FG = "FFFFFF"       # White
COLOR_ZEBRA_BG = "F2F2F2"        # Light Grey
COLOR_MATCH_HIGH = "D9EAD3"      # Soft Green Fill
COLOR_TEXT_MATCH = "274E13"      # Dark Green Text

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_FG)
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="1F4E79")
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color="595959")
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_REGULAR = Font(name="Calibri", size=11)

FILL_HEADER = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
FILL_ZEBRA = PatternFill(start_color=COLOR_ZEBRA_BG, end_color=COLOR_ZEBRA_BG, fill_type="solid")
FILL_HIGH_MATCH = PatternFill(start_color=COLOR_MATCH_HIGH, end_color=COLOR_MATCH_HIGH, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_THIN = Side(border_style="thin", color="D9D9D9")
BORDER_CELL = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)


class ExcelReportGeneratorService:
    """Service generating multi-sheet Excel reports with styling and formatting."""

    async def generate_job_search_report(
        self,
        session: AsyncSession,
        user_id: str,
        output_path: Path | str | None = None,
    ) -> bytes:
        """
        Generate a multi-worksheet Excel report for a user's job search.

        Args:
            session: Async DB session.
            user_id: User owner ID.
            output_path: Optional path to save the .xlsx file on disk.

        Returns:
            Workbook bytes.
        """
        user_repo = UserRepository(session)
        job_repo = JobRepository(session)
        app_repo = ApplicationRepository(session)
        analytics_service = AnalyticsEngineService()

        user = await user_repo.get_by_id(user_id)
        candidate_name = user.full_name if user else "Candidate"

        # Fetch Data
        jobs = await job_repo.list_active(limit=500)
        top_matches = await job_repo.get_top_matches(limit=100)
        applications = await app_repo.get_by_user(user_id, limit=200)
        dashboard_data = await analytics_service.get_dashboard_analytics(session, user_id, days=30)

        wb = openpyxl.Workbook()

        # -------------------------------------------------------------
        # Sheet 1: Executive Summary
        # -------------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        self._build_summary_sheet(ws_summary, candidate_name, dashboard_data)

        # -------------------------------------------------------------
        # Sheet 2: Top AI Matches (>= 80%)
        # -------------------------------------------------------------
        ws_top = wb.create_sheet(title="Top AI Matches")
        self._build_top_matches_sheet(ws_top, top_matches)

        # -------------------------------------------------------------
        # Sheet 3: All Discovered Jobs
        # -------------------------------------------------------------
        ws_jobs = wb.create_sheet(title="All Jobs Inventory")
        self._build_all_jobs_sheet(ws_jobs, jobs)

        # -------------------------------------------------------------
        # Sheet 4: Application Tracker
        # -------------------------------------------------------------
        ws_apps = wb.create_sheet(title="Application Tracker")
        self._build_applications_sheet(ws_apps, applications)

        # Auto-adjust column widths across all sheets
        for sheet in wb.worksheets:
            self._autofit_columns(sheet)

        # Save to buffer or file
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()

        if output_path:
            path = Path(output_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(excel_bytes)
            logger.info("excel_report_saved_to_disk", path=str(path), bytes_size=len(excel_bytes))

        logger.info("excel_report_generated", user_id=user_id, bytes_size=len(excel_bytes))
        return excel_bytes

    def _build_summary_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        candidate_name: str,
        dashboard_data: dict,
    ) -> None:
        """Build Executive Summary KPI worksheet."""
        ws.views.sheetView[0].showGridLines = True

        ws.cell(row=2, column=2, value="HuntIQ — Job Search Executive Report").font = FONT_TITLE
        now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        ws.cell(row=3, column=2, value=f"Candidate: {candidate_name} | Generated: {now_str}").font = FONT_SUBTITLE

        summary = dashboard_data.get("summary", {})

        # KPI Metric Cards
        metrics_data = [
            ("Total Jobs Discovered", summary.get("total_jobs", 0)),
            ("High Match Opportunities (>= 80%)", summary.get("high_matches", 0)),
            ("Applications Submitted", summary.get("applications", 0)),
            ("Interview Invitations", summary.get("interviews", 0)),
            ("Offers Received", summary.get("offers", 0)),
        ]

        start_row = 5
        ws.cell(row=start_row, column=2, value="Key Performance Indicator").font = FONT_HEADER
        ws.cell(row=start_row, column=2).fill = FILL_HEADER
        ws.cell(row=start_row, column=3, value="Count").font = FONT_HEADER
        ws.cell(row=start_row, column=3).fill = FILL_HEADER

        for idx, (metric, count) in enumerate(metrics_data, start=start_row + 1):
            r_cell = ws.cell(row=idx, column=2, value=metric)
            v_cell = ws.cell(row=idx, column=3, value=count)

            r_cell.font = FONT_REGULAR
            v_cell.font = FONT_BOLD
            r_cell.border = BORDER_CELL
            v_cell.border = BORDER_CELL
            v_cell.alignment = ALIGN_CENTER

            if idx % 2 == 1:
                r_cell.fill = FILL_ZEBRA
                v_cell.fill = FILL_ZEBRA

        # Top Skills Table
        top_skills = dashboard_data.get("top_skills", [])
        if top_skills:
            skill_row = start_row + len(metrics_data) + 3
            ws.cell(row=skill_row, column=2, value="Top Demanded Tech Stack Skill").font = FONT_HEADER
            ws.cell(row=skill_row, column=2).fill = FILL_HEADER
            ws.cell(row=skill_row, column=3, value="Job Postings Count").font = FONT_HEADER
            ws.cell(row=skill_row, column=3).fill = FILL_HEADER

            for s_idx, item in enumerate(top_skills[:10], start=skill_row + 1):
                s_name = item.get("skill", "") if isinstance(item, dict) else str(item)
                s_cnt = item.get("count", 0) if isinstance(item, dict) else 1

                s_cell = ws.cell(row=s_idx, column=2, value=s_name.upper())
                c_cell = ws.cell(row=s_idx, column=3, value=s_cnt)

                s_cell.font = FONT_REGULAR
                c_cell.font = FONT_REGULAR
                s_cell.border = BORDER_CELL
                c_cell.border = BORDER_CELL
                c_cell.alignment = ALIGN_CENTER

    def _build_top_matches_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, top_matches: list) -> None:
        """Build Top AI Matches worksheet with conditional formatting."""
        ws.views.sheetView[0].showGridLines = True

        headers = [
            "Rank", "Job Title", "Company", "Composite Match %",
            "Rule Score", "Embedding Score", "Location", "Remote", "URL"
        ]

        ws.row_dimensions[1].height = 24
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER

        for row_idx, job in enumerate(top_matches, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=2, value=job.title)
            ws.cell(row=row_idx, column=3, value=job.company.name if job.company else "N/A")

            score_cell = ws.cell(row=row_idx, column=4, value=round(job.match_score or 0.0, 1))
            score_cell.alignment = ALIGN_CENTER
            score_cell.font = FONT_BOLD

            ws.cell(row=row_idx, column=5, value=round(job.rule_score or 0.0, 1)).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=6, value=round(job.embedding_score or 0.0, 1)).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=7, value=job.location or "Remote")
            ws.cell(row=row_idx, column=8, value="Yes" if job.is_remote else "No").alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=9, value=job.posting_url or "")

            # Apply borders and zebra striping
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = BORDER_CELL
                if row_idx % 2 == 1:
                    cell.fill = FILL_ZEBRA

        # Conditional formatting rule for composite score >= 80
        if len(top_matches) > 0:
            rule = CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=FILL_HIGH_MATCH)
            ws.conditional_formatting.add(f"D2:D{len(top_matches) + 1}", rule)

        ws.auto_filter.ref = f"A1:I{max(len(top_matches) + 1, 2)}"

    def _build_all_jobs_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, jobs: list) -> None:
        """Build All Jobs inventory worksheet."""
        ws.views.sheetView[0].showGridLines = True
        headers = ["Job Title", "Company", "Source", "Location", "Remote", "Salary Max", "Match %", "URL"]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER

        for row_idx, job in enumerate(jobs, start=2):
            ws.cell(row=row_idx, column=1, value=job.title)
            ws.cell(row=row_idx, column=2, value=job.company.name if job.company else "N/A")
            ws.cell(row=row_idx, column=3, value=job.sources[0].source_type if job.sources else "direct")
            ws.cell(row=row_idx, column=4, value=job.location or "N/A")
            ws.cell(row=row_idx, column=5, value="Yes" if job.is_remote else "No").alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=6, value=f"${job.salary_max:,.0f}" if job.salary_max else "N/A").alignment = ALIGN_RIGHT
            ws.cell(row=row_idx, column=7, value=round(job.match_score or 0.0, 1)).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=8, value=job.posting_url or "")

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = BORDER_CELL
                if row_idx % 2 == 1:
                    cell.fill = FILL_ZEBRA

        ws.auto_filter.ref = f"A1:H{max(len(jobs) + 1, 2)}"

    def _build_applications_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, applications: list) -> None:
        """Build Application Tracker worksheet."""
        ws.views.sheetView[0].showGridLines = True
        headers = ["Job Title", "Company", "Stage", "Applied Date", "Recruiter", "Next Interview", "Offer Amount"]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER

        for row_idx, app in enumerate(applications, start=2):
            ws.cell(row=row_idx, column=1, value=app.job.title if app.job else "N/A")
            ws.cell(row=row_idx, column=2, value=app.job.company.name if (app.job and app.job.company) else "N/A")
            ws.cell(row=row_idx, column=3, value=app.current_stage.upper()).alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=4, value=app.applied_at.strftime("%Y-%m-%d") if app.applied_at else "N/A").alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=5, value=app.recruiter_name or "N/A")
            ws.cell(row=row_idx, column=6, value=app.next_interview_at.strftime("%Y-%m-%d %H:%M") if app.next_interview_at else "N/A").alignment = ALIGN_CENTER
            ws.cell(row=row_idx, column=7, value=app.offer_amount or "N/A").alignment = ALIGN_RIGHT

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = BORDER_CELL
                if row_idx % 2 == 1:
                    cell.fill = FILL_ZEBRA

        ws.auto_filter.ref = f"A1:G{max(len(applications) + 1, 2)}"

    def _autofit_columns(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Auto-adjust column widths for optimal readability."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
